#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# author: Alexey Koshevoy

from openpyxl import load_workbook, Workbook
import re
import warnings
import pymorphy2

morph = pymorphy2.MorphAnalyzer()


class Table:

    def __init__(self, path, sheet):
        self.path = path
        self.sheet_name = sheet
        self.sheet = self.get_sheet()

    @staticmethod
    def read_table(path):
        wb = load_workbook(path, read_only=False)
        return wb

    def get_sheet(self):
        """
        Make it optional – only if there is some sheets??
        """
        wb = self.read_table(path=self.path)
        sheet = wb[self.sheet_name]
        return sheet

    def get_horizontal_list(self):

        horizontal_list = []

        sheet = self.sheet
        for row in range(2, sheet.max_row):
            for column in "A":
                cell_name = "{}{}".format(column, row)
                horizontal_list.append(sheet[cell_name].value)

        return horizontal_list

    @staticmethod
    def iter_rows(ws):
        for row in ws.iter_rows():
            yield [cell.value for cell in row]

    def get_vertical_list(self):
        sheet = self.sheet
        vertical_list = list(self.iter_rows(sheet))[0]
        return vertical_list


class Translation:

    def __init__(self, list_of_words):
        """
        Maybe it is better to make a possibility to insert list of lists
        """
        self.list_of_words = list_of_words

    @staticmethod
    def get_dictionary():
        """
        Add way to add custom patch to dictionary
        :return:
        """
        path = '/Users/Alexey/Documents/automated_lt/dictionary/ru-de.txt'
        dictionary = open(path).readlines()
        return dictionary

    @staticmethod
    def translate_word(word, dictionary):
        word = re.sub(r'[^\w\s]', '', word)
        into_word = morph.parse(word)[0]
        normalized = re.sub('ё', 'е', into_word.normal_form)

        for line in dictionary:
            if re.search('^{}'.format(normalized), line):
                return re.search('([a-zA-Z0-9äöüÄÖÜ]+)', line).group(0)

    def change_list(self, dictionary):

        """
        Maybe get only words from sentence
        :param dictionary:
        :return:
        """
        list_words = self.list_of_words

        pattern = '(([А-Я]?)[а-я]*) ((\(?).*(\)?))'

        for i in range(len(list_words)):
            if list_words[i]:
                if re.match('([А-Я]?)[а-я]* ((\(?).*(\)?))', list_words[i]):
                    word1 = self.translate_word(re.search(pattern,
                                                          list_words[i]).
                                                group(1),
                                                dictionary=dictionary)
                    word2 = self.translate_word(re.search(pattern,
                                                          list_words[i]).
                                                group(3),
                                                dictionary=dictionary)
                    list_words[i] = '{} ({})'.format(word1, word2)
                else:
                    list_words[i] = self.\
                        translate_word(list_words[i], dictionary=dictionary)
        return list_words


class TranslatedTable(Table):

    def __init__(self, path, sheet, table_name=None):
        """
        What else?

        отнаследовать от Qu table?
        стуртура должна быть сложнее в таком случае (подумать)
        :param path:
        :param sheet:
        """
        Table.__init__(self, path, sheet)
        self.table_name = table_name
        self.translated = []

    def translate(self):
        list_of_two = [self.get_horizontal_list(), self.get_vertical_list()]
        print(list_of_two[0])
        for list_one in list_of_two:
            translation = Translation(list_one)
            final_trans = translation.change_list(translation.get_dictionary())
            print(final_trans)
            self.translated.append(final_trans)

    def write_translations(self):
        wb = self.read_table(self.path)
        translated_ws = wb.create_sheet('нем стандартный вид')

        r = 2

        for element in self.translated[0]:
            translated_ws.cell(row=r, column=1).value = element
            r += 1

        k = 1

        for element in self.translated[1]:
            translated_ws.cell(row=1, column=k).value = element
            k += 1

        wb.save(self.path)

if __name__ == '__main__':
    warnings.filterwarnings("ignore")
    a = TranslatedTable(path='/Users/Alexey/Documents/automated_lt/questionnaire'
                         '/questionnaire_size.xlsx',
                    sheet='русский_стандртный_вид')

    a.translate()
    a.write_translations()
